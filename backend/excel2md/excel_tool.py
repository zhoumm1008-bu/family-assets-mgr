import argparse
import logging
import os
import sys
from typing import Dict, Optional

import pandas as pd
from openpyxl import load_workbook

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

def process_worksheet(ws) -> pd.DataFrame:
    """
    处理单个工作表：解除合并、填充值、清理换行符、转为 DataFrame。
    """
    # 获取合并区域并转换为列表
    merged_ranges = list(ws.merged_cells.ranges)
    if merged_ranges:
        logger.info(f"  检测到 {len(merged_ranges)} 个合并区域，正在填充...")
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_val = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_val

    # 统一清理所有单元格中的换行符
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')

    # 转换为 DataFrame
    data = ws.values
    try:
        cols = list(next(data))
        cols = [c.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ') if isinstance(c, str) else c for c in cols]
        df = pd.DataFrame(data, columns=cols)
        # 过滤掉全为空的行
        df = df.dropna(how='all')
    except StopIteration:
        df = pd.DataFrame()
    return df

def parse_excel_to_dfs(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, pd.DataFrame]:
    """
    加载 Excel 并处理指定或所有工作表。
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"找不到文件: {file_path}")

    logger.info(f"正在读取 Excel 文件: {file_path}")
    wb = load_workbook(file_path, data_only=True)
    
    results = {}
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"找不到工作表 '{sheet_name}'。可用: {wb.sheetnames}")
        sheets_to_process = [sheet_name]
    else:
        sheets_to_process = wb.sheetnames
        logger.info(f"未指定工作表，将自动处理所有 {len(sheets_to_process)} 个工作表")

    for name in sheets_to_process:
        logger.info(f"正在处理工作表: {name}...")
        results[name] = process_worksheet(wb[name])
    
    return results

def main():
    parser = argparse.ArgumentParser(description="Excel 转 CSV 工具（支持多 Sheet 及合并单元格处理）")
    parser.add_argument("input", help="输入的 Excel 文件路径 (.xlsx)")
    parser.add_argument("-o", "--output", help="输出路径 (单表时为文件名，多表时为前缀)")
    parser.add_argument("-s", "--sheet", help="指定工作表名称 (缺省则转换所有)")
    parser.add_argument("--encoding", default="utf-8-sig", help="CSV 编码格式 (默认 utf-8-sig)")

    args = parser.parse_args()
    input_base = os.path.splitext(args.input)[0]

    try:
        results = parse_excel_to_dfs(args.input, args.sheet)
        
        for sheet_name, df in results.items():
            if df.empty:
                logger.warning(f"工作表 '{sheet_name}' 为空，跳过。")
                continue
            
            # 确定输出文件名
            if args.sheet:
                # 如果用户指定了特定 Sheet，优先使用 -o，否则使用输入文件名作为 base
                out_path = args.output if args.output else f"{input_base}.csv"
            else:
                # 如果是自动转换所有 Sheet，文件名加上 Sheet 名
                out_path = f"{input_base}_{sheet_name}.csv"
                
            df.to_csv(out_path, index=False, encoding=args.encoding)
            logger.info(f"成功保存到: {out_path}")
            
    except Exception as e:
        logger.error(f"处理失败: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
